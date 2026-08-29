"""
Finesse x Citadel Portfolio Challenge -- Round 2
Team: Anti Buldhana Gang  |  Members: Raghav Fulara, Mayuresh Thengzode, Nirmit Bangar

Buffered momentum strategy ("BMOM 12-1") for a 10-stock, Rs 1 crore
long-only portfolio over 1 Jan 2021 - 31 Dec 2025.

The idea is simple: every quarter, rank the full eligible universe
(Nifty 100 + Midcap 100 + Smallcap 100 constituents, ~300 stocks) by
12-month momentum measured with a 1-month skip, and hold the strongest
names. Two details do most of the work:

  * the 1-month skip -- recent winners tend to reverse in India, so the
    signal looks at t-273 .. t-21 instead of the last 12 months directly
  * a rank buffer -- we keep holding a stock while it stays inside the
    top 20 by signal rank, and only replace it once it drops out. This
    keeps turnover (and hence cost + tax drag) low and lets the long
    multi-year trends actually compound.

Weights are inverse 60-day volatility with a 20% cap per name, and a
30% stop-loss versus the last review price protects against blow-ups.
0.1% transaction cost is charged on every trade, per the problem stat.
ment. The same frozen rules are used for the 2021-25 backtest and the
Jan-Jun 2026 out-of-sample run -- nothing is re-fitted anywhere.

Note: iterations/ holds earlier fractional-share prototypes of this
strategy (v1-v3); this version re-implements it with share-level
accounting (integer lots, cash ledger) and freezes the stop at 30%
after probing the 25/30 plateau, so its numbers differ slightly from
the prototypes.

Run:  python final_strategy.py     (~2 min with the cached price file)
"""


import io
import json
import contextlib
from datetime import timedelta

import numpy as np
import pandas as pd
import yfinance as yf
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats

import warnings
warnings.filterwarnings('ignore')

plt.style.use('seaborn-v0_8-darkgrid')
np.random.seed(42)

# =====================================================================
#  CONFIG - all strategy choices in one place.
#  (identical for the 2021-25 backtest and the 2026 out-of-sample run)
# =====================================================================
CAPITAL    = 10_000_000          # ₹1 crore
TXN_COST   = 0.001               # 0.1% per transaction (per SEBI guideline)
IS_START   = '2021-01-01'
IS_END     = '2025-12-31'        # official backtest window
OOS_START  = '2026-01-01'
OOS_END    = '2026-06-30'        # official out-of-sample stress window
DATA_START = '2019-01-01'        # extra history for momentum lookback

CONFIG = dict(
    n_stocks   = 7,       # portfolio size (≤ 10 allowed)
    keep_rank  = 20,      # rank buffer: hold while rank < keep_rank
    mom_lb     = 252,     # momentum lookback (12 months)
    mom_skip   = 21,      # skip most recent month (reversal guard)
    vol_window = 60,      # vol window for inverse-vol weights
    max_w      = 0.20,    # per-stock weight cap
    rebal      = 'Q',     # 'Q' quarterly | 'M' monthly | '2M'
    stop_loss  = -0.30,   # catastrophic stop vs review price (None = off; -0.25/-0.30 plateau)
    min_mom    = 0.0,     # absolute-momentum floor (only mom > 0)
    ind_cap    = None,    # optional: max NEW entries per NSE industry (None = off; costs ~1 Cr)
)

# =====================================================================
#  UNIVERSE - official constituents (Nifty 100 + Midcap 100 + Smallcap
#  100, union = 300 names, pulled from niftyindices.com on 2026-08-22).
#  refresh_universe() re-pulls the latest official lists if reachable.
# =====================================================================
UNIVERSE = [
    '360ONE', 'AARTIIND', 'ABB', 'ABCAPITAL', 'ABREL', 'ADANIENSOL', 'ADANIENT',
    'ADANIGREEN', 'ADANIPORTS', 'ADANIPOWER', 'AEGISLOG', 'AFCONS', 'AFFLE', 'ALKEM',
    'AMBER', 'AMBUJACEM', 'ANANDRATHI', 'ANANTRAJ', 'ANGELONE', 'APLAPOLLO', 'APOLLOHOSP',
    'APTUS', 'ARE&M', 'ASHOKLEY', 'ASIANPAINT', 'ASTERDM', 'ASTRAL', 'ATGL', 'ATHERENERG',
    'AUBANK', 'AUROPHARMA', 'AXISBANK', 'BAJAJ-AUTO', 'BAJAJFINSV', 'BAJAJHLDNG',
    'BAJFINANCE', 'BANDHANBNK', 'BANKBARODA', 'BANKINDIA', 'BDL', 'BEL', 'BEML',
    'BHARATFORG', 'BHARTIARTL', 'BHEL', 'BIOCON', 'BLS', 'BLUESTARCO', 'BOSCHLTD', 'BPCL',
    'BRIGADE', 'BRITANNIA', 'BSE', 'CAMS', 'CANBK', 'CASTROLIND', 'CDSL', 'CESC', 'CGCL',
    'CGPOWER', 'CHAMBLFERT', 'CHOLAFIN', 'CHOLAHLDNG', 'CIPLA', 'COALINDIA', 'COCHINSHIP',
    'COFORGE', 'COHANCE', 'COLPAL', 'CONCOR', 'COROMANDEL', 'CREDITACC', 'CROMPTON', 'CUB',
    'CUMMINSIND', 'DABUR', 'DATAPATTNS', 'DEEPAKFERT', 'DELHIVERY', 'DEVYANI', 'DIVISLAB',
    'DIXON', 'DLF', 'DMART', 'DRREDDY', 'EICHERMOT', 'ENRIN', 'ETERNAL', 'EXIDEIND',
    'FEDERALBNK', 'FIRSTCRY', 'FIVESTAR', 'FORCEMOT', 'FORTIS', 'FSL', 'GAIL', 'GESHIP',
    'GLAND', 'GLENMARK', 'GMDCLTD', 'GMRAIRPORT', 'GODFRYPHLP', 'GODREJCP', 'GODREJPROP',
    'GPIL', 'GRASIM', 'GROWW', 'GRSE', 'GVT&D', 'HAL', 'HAVELLS', 'HBLENGINE', 'HCLTECH',
    'HDFCAMC', 'HDFCBANK', 'HDFCLIFE', 'HEROMOTOCO', 'HINDALCO', 'HINDCOPPER', 'HINDPETRO',
    'HINDUNILVR', 'HINDZINC', 'HSCL', 'HUDCO', 'HYUNDAI', 'ICICIAMC', 'ICICIBANK', 'ICICIGI',
    'IDBI', 'IDEA', 'IDFCFIRSTB', 'IFCI', 'IGL', 'IIFL', 'IKS', 'INDHOTEL', 'INDIANB',
    'INDIGO', 'INDUSINDBK', 'INDUSTOWER', 'INFY', 'INOXWIND', 'IOC', 'IRCON', 'IRCTC',
    'IREDA', 'IRFC', 'ITC', 'ITI', 'JBMA', 'JINDALSTEL', 'JIOFIN', 'JMFINANCIL', 'JSWCEMENT',
    'JSWENERGY', 'JSWSTEEL', 'JUBLFOOD', 'JYOTICNC', 'KALYANKJIL', 'KARURVYSYA', 'KAYNES',
    'KEC', 'KEI', 'KFINTECH', 'KOTAKBANK', 'KPITTECH', 'LALPATHLAB', 'LAURUSLABS',
    'LENSKART', 'LGEINDIA', 'LICHSGFIN', 'LODHA', 'LT', 'LTF', 'LTM', 'LUPIN', 'M&M',
    'M&MFIN', 'MANAPPURAM', 'MANKIND', 'MARICO', 'MARUTI', 'MAXHEALTH', 'MAZDOCK', 'MCX',
    'MEESHO', 'MFSL', 'MOTHERSON', 'MOTILALOFS', 'MPHASIS', 'MRF', 'MRPL', 'MUTHOOTFIN',
    'NATCOPHARM', 'NATIONALUM', 'NAUKRI', 'NAVINFLUOR', 'NBCC', 'NESTLEIND', 'NETWEB',
    'NEULANDLAB', 'NH', 'NHPC', 'NMDC', 'NTPC', 'NUVAMA', 'NYKAA', 'OBEROIRLTY', 'OFSS',
    'OIL', 'OLAELEC', 'ONGC', 'PAGEIND', 'PATANJALI', 'PAYTM', 'PERSISTENT', 'PFC', 'PGEL',
    'PHOENIXLTD', 'PIDILITIND', 'PIIND', 'PINELABS', 'PIRAMALFIN', 'PNB', 'PNBHOUSING',
    'POLICYBZR', 'POLYCAB', 'POONAWALLA', 'POWERGRID', 'POWERINDIA', 'PPLPHARMA',
    'PREMIERENE', 'PRESTIGE', 'PWL', 'RADICO', 'RAMCOCEM', 'RBLBANK', 'RECLTD', 'REDINGTON',
    'RELIANCE', 'RPOWER', 'RVNL', 'SAGILITY', 'SAIL', 'SAILIFE', 'SARDAEN', 'SBICARD',
    'SBILIFE', 'SBIN', 'SHREECEM', 'SHRIRAMFIN', 'SIEMENS', 'SIGNATURE', 'SOLARINDS',
    'SONACOMS', 'SRF', 'STARHEALTH', 'SUNPHARMA', 'SUPREMEIND', 'SUZLON', 'SWANCORP',
    'SWIGGY', 'SYNGENE', 'TATACAP', 'TATACHEM', 'TATACOMM', 'TATACONSUM', 'TATAELXSI',
    'TATAINVEST', 'TATAPOWER', 'TATASTEEL', 'TATATECH', 'TCS', 'TECHM', 'TENNIND', 'TIINDIA',
    'TITAN', 'TMCV', 'TMPV', 'TORNTPHARM', 'TRENT', 'TRITURBINE', 'TVSMOTOR', 'ULTRACEMCO',
    'UNIONBANK', 'UNITDSPR', 'UPL', 'URBANCO', 'VBL', 'VEDL', 'VMM', 'VOLTAS', 'WAAREEENER',
    'WELCORP', 'WHIRLPOOL', 'WIPRO', 'WOCKPHARMA', 'YESBANK', 'ZENSARTECH', 'ZYDUSLIFE',
]


# NSE industry classification (official constituent files, niftyindices.com)
INDUSTRY = {
    'ABB': 'Capital Goods', 'ADANIENSOL': 'Power', 'ADANIENT': 'Metals & Mining',
    'ADANIGREEN': 'Power', 'ADANIPORTS': 'Services', 'ADANIPOWER': 'Power',
    'AMBUJACEM': 'Construction Materials', 'APOLLOHOSP': 'Healthcare',
    'ASIANPAINT': 'Consumer Durables', 'DMART': 'Consumer Services',
    'AXISBANK': 'Financial Services', 'BAJAJ-AUTO': 'Automobile and Auto Components',
    'BAJFINANCE': 'Financial Services', 'BAJAJFINSV': 'Financial Services',
    'BAJAJHLDNG': 'Financial Services', 'BANKBARODA': 'Financial Services',
    'BEL': 'Capital Goods', 'BPCL': 'Oil Gas & Consumable Fuels',
    'BHARTIARTL': 'Telecommunication', 'BOSCHLTD': 'Automobile and Auto Components',
    'BRITANNIA': 'Fast Moving Consumer Goods', 'CGPOWER': 'Capital Goods',
    'CANBK': 'Financial Services', 'CHOLAFIN': 'Financial Services', 'CIPLA': 'Healthcare',
    'COALINDIA': 'Oil Gas & Consumable Fuels', 'CUMMINSIND': 'Capital Goods', 'DLF': 'Realty',
    'DIVISLAB': 'Healthcare', 'DRREDDY': 'Healthcare',
    'EICHERMOT': 'Automobile and Auto Components', 'ETERNAL': 'Consumer Services',
    'GAIL': 'Oil Gas & Consumable Fuels', 'GODREJCP': 'Fast Moving Consumer Goods',
    'GRASIM': 'Construction Materials', 'HCLTECH': 'Information Technology',
    'HDFCAMC': 'Financial Services', 'HDFCBANK': 'Financial Services',
    'HDFCLIFE': 'Financial Services', 'HINDALCO': 'Metals & Mining', 'HAL': 'Capital Goods',
    'HINDUNILVR': 'Fast Moving Consumer Goods', 'HINDZINC': 'Metals & Mining',
    'HYUNDAI': 'Automobile and Auto Components', 'ICICIBANK': 'Financial Services',
    'ITC': 'Fast Moving Consumer Goods', 'INDHOTEL': 'Consumer Services',
    'IOC': 'Oil Gas & Consumable Fuels', 'IRFC': 'Financial Services',
    'INFY': 'Information Technology', 'INDIGO': 'Services', 'JSWSTEEL': 'Metals & Mining',
    'JINDALSTEL': 'Metals & Mining', 'JIOFIN': 'Financial Services',
    'KOTAKBANK': 'Financial Services', 'LTM': 'Information Technology', 'LT': 'Construction',
    'LODHA': 'Realty', 'M&M': 'Automobile and Auto Components',
    'MARUTI': 'Automobile and Auto Components', 'MAXHEALTH': 'Healthcare',
    'MAZDOCK': 'Capital Goods', 'MUTHOOTFIN': 'Financial Services', 'NTPC': 'Power',
    'NESTLEIND': 'Fast Moving Consumer Goods', 'ONGC': 'Oil Gas & Consumable Fuels',
    'PIDILITIND': 'Chemicals', 'PFC': 'Financial Services', 'POWERGRID': 'Power',
    'PNB': 'Financial Services', 'RECLTD': 'Financial Services',
    'RELIANCE': 'Oil Gas & Consumable Fuels', 'SBILIFE': 'Financial Services',
    'MOTHERSON': 'Automobile and Auto Components', 'SHREECEM': 'Construction Materials',
    'SHRIRAMFIN': 'Financial Services', 'ENRIN': 'Capital Goods', 'SIEMENS': 'Capital Goods',
    'SOLARINDS': 'Chemicals', 'SBIN': 'Financial Services', 'SUNPHARMA': 'Healthcare',
    'TVSMOTOR': 'Automobile and Auto Components', 'TATACAP': 'Financial Services',
    'TCS': 'Information Technology', 'TATACONSUM': 'Fast Moving Consumer Goods',
    'TMCV': 'Capital Goods', 'TMPV': 'Automobile and Auto Components', 'TATAPOWER': 'Power',
    'TATASTEEL': 'Metals & Mining', 'TECHM': 'Information Technology',
    'TITAN': 'Consumer Durables', 'TORNTPHARM': 'Healthcare', 'TRENT': 'Consumer Services',
    'ULTRACEMCO': 'Construction Materials', 'UNIONBANK': 'Financial Services',
    'UNITDSPR': 'Fast Moving Consumer Goods', 'VBL': 'Fast Moving Consumer Goods',
    'VEDL': 'Metals & Mining', 'WIPRO': 'Information Technology', 'ZYDUSLIFE': 'Healthcare',
    '360ONE': 'Financial Services', 'APLAPOLLO': 'Capital Goods',
    'AUBANK': 'Financial Services', 'ATGL': 'Oil Gas & Consumable Fuels',
    'ABCAPITAL': 'Financial Services', 'ALKEM': 'Healthcare', 'ASHOKLEY': 'Capital Goods',
    'ASTRAL': 'Capital Goods', 'AUROPHARMA': 'Healthcare', 'BSE': 'Financial Services',
    'BANKINDIA': 'Financial Services', 'BDL': 'Capital Goods',
    'BHARATFORG': 'Automobile and Auto Components', 'BHEL': 'Capital Goods',
    'GROWW': 'Financial Services', 'BIOCON': 'Healthcare', 'BLUESTARCO': 'Consumer Durables',
    'COCHINSHIP': 'Capital Goods', 'COFORGE': 'Information Technology',
    'COLPAL': 'Fast Moving Consumer Goods', 'CONCOR': 'Services', 'COROMANDEL': 'Chemicals',
    'DABUR': 'Fast Moving Consumer Goods', 'DIXON': 'Consumer Durables',
    'EXIDEIND': 'Automobile and Auto Components', 'NYKAA': 'Consumer Services',
    'FEDERALBNK': 'Financial Services', 'FORTIS': 'Healthcare', 'GVT&D': 'Capital Goods',
    'GMRAIRPORT': 'Services', 'GLENMARK': 'Healthcare',
    'GODFRYPHLP': 'Fast Moving Consumer Goods', 'GODREJPROP': 'Realty',
    'HAVELLS': 'Consumer Durables', 'HEROMOTOCO': 'Automobile and Auto Components',
    'HINDPETRO': 'Oil Gas & Consumable Fuels', 'POWERINDIA': 'Capital Goods',
    'HUDCO': 'Financial Services', 'ICICIGI': 'Financial Services',
    'ICICIAMC': 'Financial Services', 'IDFCFIRSTB': 'Financial Services',
    'INDIANB': 'Financial Services', 'IRCTC': 'Consumer Services',
    'IREDA': 'Financial Services', 'INDUSTOWER': 'Telecommunication',
    'INDUSINDBK': 'Financial Services', 'NAUKRI': 'Consumer Services', 'JSWENERGY': 'Power',
    'JUBLFOOD': 'Consumer Services', 'KEI': 'Capital Goods',
    'KPITTECH': 'Information Technology', 'KALYANKJIL': 'Consumer Durables',
    'LTF': 'Financial Services', 'LGEINDIA': 'Consumer Durables',
    'LICHSGFIN': 'Financial Services', 'LAURUSLABS': 'Healthcare',
    'LENSKART': 'Consumer Services', 'LUPIN': 'Healthcare',
    'MRF': 'Automobile and Auto Components', 'M&MFIN': 'Financial Services',
    'MANKIND': 'Healthcare', 'MARICO': 'Fast Moving Consumer Goods',
    'MFSL': 'Financial Services', 'MOTILALOFS': 'Financial Services',
    'MPHASIS': 'Information Technology', 'MCX': 'Financial Services', 'NHPC': 'Power',
    'NMDC': 'Metals & Mining', 'NATIONALUM': 'Metals & Mining', 'OBEROIRLTY': 'Realty',
    'OIL': 'Oil Gas & Consumable Fuels', 'PAYTM': 'Financial Services',
    'OFSS': 'Information Technology', 'POLICYBZR': 'Financial Services', 'PIIND': 'Chemicals',
    'PAGEIND': 'Textiles', 'PATANJALI': 'Fast Moving Consumer Goods',
    'PERSISTENT': 'Information Technology', 'PHOENIXLTD': 'Realty', 'POLYCAB': 'Capital Goods',
    'PREMIERENE': 'Capital Goods', 'PRESTIGE': 'Realty',
    'RADICO': 'Fast Moving Consumer Goods', 'RVNL': 'Construction',
    'SBICARD': 'Financial Services', 'SRF': 'Chemicals', 'SAIL': 'Metals & Mining',
    'SUPREMEIND': 'Capital Goods', 'SUZLON': 'Capital Goods', 'SWIGGY': 'Consumer Services',
    'TATACOMM': 'Telecommunication', 'TATAELXSI': 'Information Technology',
    'TATAINVEST': 'Financial Services', 'TIINDIA': 'Automobile and Auto Components',
    'UPL': 'Chemicals', 'VMM': 'Consumer Services', 'IDEA': 'Telecommunication',
    'VOLTAS': 'Consumer Durables', 'WAAREEENER': 'Capital Goods',
    'YESBANK': 'Financial Services', 'AARTIIND': 'Chemicals', 'ABREL': 'Realty',
    'AEGISLOG': 'Oil Gas & Consumable Fuels', 'AFCONS': 'Construction',
    'AFFLE': 'Information Technology', 'ARE&M': 'Automobile and Auto Components',
    'AMBER': 'Consumer Durables', 'ANANDRATHI': 'Financial Services', 'ANANTRAJ': 'Realty',
    'ANGELONE': 'Financial Services', 'APTUS': 'Financial Services', 'ASTERDM': 'Healthcare',
    'ATHERENERG': 'Automobile and Auto Components', 'BEML': 'Capital Goods',
    'BLS': 'Consumer Services', 'BANDHANBNK': 'Financial Services',
    'FIRSTCRY': 'Consumer Services', 'BRIGADE': 'Realty', 'CESC': 'Power',
    'CGCL': 'Financial Services', 'CASTROLIND': 'Oil Gas & Consumable Fuels',
    'CDSL': 'Financial Services', 'CHAMBLFERT': 'Chemicals',
    'CHOLAHLDNG': 'Financial Services', 'CUB': 'Financial Services', 'COHANCE': 'Healthcare',
    'CAMS': 'Financial Services', 'CREDITACC': 'Financial Services',
    'CROMPTON': 'Consumer Durables', 'DATAPATTNS': 'Capital Goods', 'DEEPAKFERT': 'Chemicals',
    'DELHIVERY': 'Services', 'DEVYANI': 'Consumer Services', 'LALPATHLAB': 'Healthcare',
    'FSL': 'Services', 'FIVESTAR': 'Financial Services',
    'FORCEMOT': 'Automobile and Auto Components', 'GRSE': 'Capital Goods',
    'GLAND': 'Healthcare', 'GPIL': 'Capital Goods', 'GESHIP': 'Services',
    'GMDCLTD': 'Metals & Mining', 'HBLENGINE': 'Capital Goods', 'HSCL': 'Chemicals',
    'HINDCOPPER': 'Metals & Mining', 'IDBI': 'Financial Services',
    'IFCI': 'Financial Services', 'IIFL': 'Financial Services', 'IRCON': 'Construction',
    'ITI': 'Telecommunication', 'IGL': 'Oil Gas & Consumable Fuels',
    'INOXWIND': 'Capital Goods', 'IKS': 'Information Technology',
    'JBMA': 'Automobile and Auto Components', 'JMFINANCIL': 'Financial Services',
    'JSWCEMENT': 'Construction Materials', 'JYOTICNC': 'Capital Goods',
    'KARURVYSYA': 'Financial Services', 'KAYNES': 'Capital Goods', 'KEC': 'Construction',
    'KFINTECH': 'Financial Services', 'MANAPPURAM': 'Financial Services',
    'MRPL': 'Oil Gas & Consumable Fuels', 'MEESHO': 'Consumer Services',
    'NATCOPHARM': 'Healthcare', 'NBCC': 'Construction', 'NH': 'Healthcare',
    'NAVINFLUOR': 'Chemicals', 'NETWEB': 'Information Technology', 'NEULANDLAB': 'Healthcare',
    'NUVAMA': 'Financial Services', 'OLAELEC': 'Automobile and Auto Components',
    'PGEL': 'Consumer Durables', 'PNBHOUSING': 'Financial Services',
    'PWL': 'Consumer Services', 'PINELABS': 'Financial Services',
    'PIRAMALFIN': 'Financial Services', 'PPLPHARMA': 'Healthcare',
    'POONAWALLA': 'Financial Services', 'RBLBANK': 'Financial Services',
    'REDINGTON': 'Services', 'RPOWER': 'Power', 'SAGILITY': 'Information Technology',
    'SAILIFE': 'Healthcare', 'SARDAEN': 'Metals & Mining', 'SIGNATURE': 'Realty',
    'SONACOMS': 'Automobile and Auto Components', 'STARHEALTH': 'Financial Services',
    'SWANCORP': 'Chemicals', 'SYNGENE': 'Healthcare', 'TATACHEM': 'Chemicals',
    'TATATECH': 'Information Technology', 'TENNIND': 'Automobile and Auto Components',
    'RAMCOCEM': 'Construction Materials', 'TRITURBINE': 'Capital Goods',
    'URBANCO': 'Consumer Services', 'WELCORP': 'Capital Goods',
    'WHIRLPOOL': 'Consumer Durables', 'WOCKPHARMA': 'Healthcare',
    'ZENSARTECH': 'Information Technology',
}


def make_universe():
    return sorted(set(UNIVERSE))


def refresh_universe(path='universe_official.txt'):
    """Optionally re-pull the live official constituent lists."""
    import requests
    hdrs = {'User-Agent': 'Mozilla/5.0', 'Referer': 'https://www.niftyindices.com/'}
    out = set()
    for idx in ['nifty100', 'niftymidcap100', 'niftysmallcap100']:
        url = f'https://www.niftyindices.com/IndexConstituent/ind_{idx}list.csv'
        r = requests.get(url, headers=hdrs, timeout=20)
        if r.status_code == 200 and 'Symbol' in r.text[:800]:
            out |= set(pd.read_csv(io.StringIO(r.text))['Symbol'])
    if len(out) > 250:
        open(path, 'w').write('\n'.join(sorted(out)))
        print(f'  refreshed universe -> {len(out)} symbols')
        return sorted(out)
    print('  refresh failed; using embedded list')
    return make_universe()


# =====================================================================
#  DATA
# =====================================================================
def download_data(start=DATA_START, end=OOS_END, use_cache=True):
    """Adjusted daily closes for the whole universe (cached to parquet).

    Note: yfinance treats `end` as EXCLUSIVE, so we pad it by a week --
    otherwise the last trading day of the evaluation window (30 Jun 2026)
    silently drops out of the data.
    """
    import os
    cache = 'prices_full.parquet'
    if use_cache and os.path.exists(cache):
        px = pd.read_parquet(cache)
        px.index = pd.to_datetime(px.index)
        print(f'  loaded cached prices: {px.shape}')
        return px
    syms = [s + '.NS' for s in make_universe()]
    fetch_end = (pd.Timestamp(end) + pd.Timedelta(days=7)).strftime('%Y-%m-%d')
    print(f'  downloading {len(syms)} symbols …')
    raw = yf.download(syms, start=start, end=fetch_end, auto_adjust=True,
                      progress=False, threads=True)
    close = raw.xs('Close', axis=1, level='Price')
    close = close.loc[:, ~close.columns.duplicated()]
    # Yahoo occasionally drops a ticker due to rate limiting; retry those once
    dead = [c for c in close.columns if close[c].notna().sum() == 0]
    if dead:
        print(f'  retrying {len(dead)} dropped ticker(s): '
              f'{[c.replace(".NS", "") for c in dead]}')
        for s_ in dead:
            try:
                d = yf.download(s_, start=start, end=fetch_end,
                                auto_adjust=True, progress=False)
                if not d.empty:
                    close[s_] = d['Close'].squeeze().reindex(close.index)
            except Exception:
                pass
        still = [c for c in close.columns if close[c].notna().sum() == 0]
        if still:
            print(f'  WARNING: no data for {still} (continuing without them)')
    close = close.ffill(limit=5)          # bridge ≤5-day data gaps only
    close.to_parquet(cache)
    print(f'  data shape: {close.shape}')
    return close


def download_benchmarks(start=DATA_START, end=OOS_END):
    out = {}
    fetch_end = (pd.Timestamp(end) + pd.Timedelta(days=7)).strftime('%Y-%m-%d')
    for name, tkr in [('Nifty 50', '^NSEI'), ('Nifty 500', '^CRSLDX'),
                      ('Nifty Midcap 50', '^NSEMDCP50')]:
        try:
            d = yf.download(tkr, start=start, end=fetch_end, auto_adjust=True,
                            progress=False)
            if not d.empty:
                s = d['Close'].squeeze()
                out[name] = s
        except Exception:
            pass
    return out


# =====================================================================
#  SIGNAL -- computed strictly from data up to and including t
#      (no lookahead: a review at date t never sees beyond t)
# =====================================================================
def momentum_scores(close, t, lb=None, skip=None):
    lb = lb or CONFIG['mom_lb']
    skip = skip or CONFIG['mom_skip']
    hist = close.loc[:t]
    if len(hist) < lb + skip + 5:
        return pd.Series(dtype=float)
    p_now = hist.iloc[-1 - skip]
    p_then = hist.iloc[-1 - skip - lb]
    return (p_now / p_then - 1.0)


class BufferedMomentumSelector:
    """Pick the portfolio at a review date.

    Remembers which symbols we currently hold, so a position is only
    dropped once it falls past `keep_rank` in the signal ranking (the
    "buffer"). Empty slots are filled with the best-ranked newcomers.
    """

    def __init__(self, cfg=None):
        self.cfg = cfg or CONFIG
        self.current = []

    def reset(self):
        self.current = []

    def __call__(self, close, t):
        cfg = self.cfg
        score = momentum_scores(close, t, cfg['mom_lb'], cfg['mom_skip'])
        if score.empty:
            return {}
        score = score[score > cfg['min_mom']].dropna()
        score = score.sort_values(ascending=False)
        rank = {s: i for i, s in enumerate(score.index)}

        # keep incumbents while still inside the buffer (best-ranked first)
        keep = sorted([s for s in self.current
                       if s in rank and rank[s] < cfg['keep_rank']],
                      key=lambda s: rank[s])[:cfg['n_stocks']]
        # fill remaining slots with best-ranked newcomers,
        # respecting the per-industry cap on new entries
        def ind_ok(sel, s):
            cap = cfg.get('ind_cap')
            if not cap:
                return True
            counts = {}
            for x in sel:
                counts[INDUSTRY.get(x.replace('.NS', ''), '?')] = \
                    counts.get(INDUSTRY.get(x.replace('.NS', ''), '?'), 0) + 1
            return counts.get(INDUSTRY.get(s.replace('.NS', ''), '?'), 0) < cap
        for s in score.index:
            if len(keep) >= cfg['n_stocks']:
                break
            if s not in keep and ind_ok(keep, s):
                keep.append(s)
        self.current = list(keep)

        sel = score.reindex(keep).dropna()
        if sel.empty:
            return {}

        hist = close.loc[:t]
        vol = (hist.iloc[-cfg['vol_window']:].pct_change().std()
               * np.sqrt(252)).reindex(sel.index)
        inv = 1.0 / vol.clip(lower=0.05)
        w = inv / inv.sum()
        w = w.clip(upper=cfg['max_w'])
        w = w / w.sum()
        return w.to_dict()


# =====================================================================
#  BACKTESTER
#  Share-level accounting: integer lots, cash ledger, 0.1% cost on
#  every fill, daily stop-loss checks, quarterly reviews.
# =====================================================================
class Backtester:
    def __init__(self, close_df, start, end, capital=CAPITAL,
                 txn_cost=TXN_COST, cfg=None):
        self.close = close_df
        self.start = pd.to_datetime(start)
        self.end = pd.to_datetime(end)
        self.capital0 = float(capital)
        self.cash = float(capital)
        self.txn_cost = txn_cost
        self.cfg = dict(cfg or CONFIG)

        self.holdings = {}        # symbol -> shares
        self.review_px = {}       # symbol -> price at last review (for stop)
        self.trades = []
        self.nav_series = {}
        self.selector = BufferedMomentumSelector(self.cfg)
        self.selector_log = []   # [(date, {symbol: weight})]
        self.episodes = {}       # symbol -> {'shares','cost','realized'}

    # latest available price at or before `date` for a symbol
    # (carries the last traded price through data gaps)
    def _px(self, sym, date):
        try:
            s = self.close[sym].loc[:date].dropna()
            return float(s.iloc[-1]) if not s.empty else np.nan
        except Exception:
            return np.nan

    def _value(self, date):
        v = self.cash
        for sym, sh in self.holdings.items():
            px = self._px(sym, date)
            if not np.isnan(px):
                v += sh * px
        return v

    # -- trade primitives --
    def _buy(self, sym, cash_amt, date):
        px = self._px(sym, date)
        if np.isnan(px) or cash_amt <= 0 or px <= 0:
            return 0
        sh = int(cash_amt / (px * (1 + self.txn_cost)))
        if sh <= 0:
            return 0
        gross = sh * px
        cost = gross * self.txn_cost
        self.cash -= (gross + cost)
        self.holdings[sym] = self.holdings.get(sym, 0) + sh
        ep = self.episodes.setdefault(sym, {'shares': 0, 'cost': 0.0, 'realized': 0.0})
        ep['shares'] += sh
        ep['cost'] += gross + cost
        self.trades.append({'date': date, 'symbol': sym, 'action': 'BUY',
                            'reason': 'REVIEW', 'shares': sh, 'price': px,
                            'gross': gross, 'cost': cost})
        return sh

    def _sell(self, sym, sh, date, reason='REVIEW'):
        px = self._px(sym, date)
        if np.isnan(px) or sh <= 0:
            return 0.0
        sh = min(sh, self.holdings.get(sym, 0))
        if sh <= 0:
            return 0.0
        gross = sh * px
        cost = gross * self.txn_cost
        self.cash += gross - cost
        self.holdings[sym] = self.holdings.get(sym, 0) - sh
        if self.holdings[sym] <= 0:
            self.holdings.pop(sym, None)
        ep = self.episodes.get(sym, {'shares': 0, 'cost': 0.0, 'realized': 0.0})
        if ep['shares'] > 0:
            avg_cost = ep['cost'] / ep['shares']
            ep['realized'] += gross - cost - avg_cost * sh
            ep['shares'] -= sh
            ep['cost'] -= avg_cost * sh
        self.trades.append({'date': date, 'symbol': sym, 'action': 'SELL',
                            'reason': reason, 'shares': sh, 'price': px,
                            'gross': gross, 'cost': cost})
        return gross - cost

    # -- risk --
    def _check_stop_loss(self, date):
        if self.cfg.get('stop_loss') is None:
            return
        for sym, sh in list(self.holdings.items()):
            if sh <= 0:
                continue
            rp = self.review_px.get(sym, np.nan)
            if np.isnan(rp):
                continue
            px = self._px(sym, date)
            if np.isnan(px):
                continue
            if px / rp - 1.0 <= self.cfg['stop_loss']:
                self._sell(sym, sh, date, reason='STOP_LOSS')
                self.review_px.pop(sym, None)

    # -- rebalance --
    def _rebalance(self, date, targets):
        nav = self._value(date)
        self.selector_log.append((date, targets))
        # 1) exit names that left the portfolio
        for sym in list(self.holdings.keys()):
            if sym not in targets:
                self._sell(sym, self.holdings[sym], date)
                self.review_px.pop(sym, None)
        # 2) size positions to targets (trim overweights first)
        for sym, w in sorted(targets.items(), key=lambda kv: kv[1]):
            px = self._px(sym, date)
            if np.isnan(px):
                continue
            tgt_val = nav * w
            cur_val = self.holdings.get(sym, 0) * px
            if cur_val - tgt_val > px:                      # trim
                sh = int((cur_val - tgt_val) / px)
                if sh > 0:
                    self._sell(sym, sh, date)
            elif tgt_val - cur_val > px:                    # top-up / open
                shortfall = tgt_val - cur_val
                self._buy(sym, min(shortfall, self.cash), date)
            self.review_px[sym] = px

    # -- main loop --
    def run(self, verbose=True):
        dates = self.close.loc[self.start:self.end].index
        if len(dates) == 0:
            print('  no dates in range')
            return self

        reb = self.cfg['rebal']
        reb_dates = set()
        prev = None
        for d in dates:
            key = (d.year, {'Q': (d.month - 1) // 3,
                            'M': d.month,
                            '2M': (d.month - 1) // 2}[reb])
            if prev is None or key != prev:
                reb_dates.add(d)
            prev = key

        for d in dates:
            self._check_stop_loss(d)
            if d in reb_dates:
                w = self.selector(self.close, d)
                if w:
                    self._rebalance(d, w)
                    if verbose:
                        print(f"  REVIEW {d.date()}: "
                              f"{[s.replace('.NS','') for s in self.holdings]}")
            self.nav_series[d] = self._value(d)
        return self


# =====================================================================
#  FAST WEIGHT-SPACE ENGINE (for stress modules only)
# =====================================================================
def fast_backtest(close, start, end, cfg=None, capital=CAPITAL, cost=TXN_COST):
    cfg = dict(cfg or CONFIG)
    sel = BufferedMomentumSelector(cfg)
    rets = close.pct_change(fill_method=None)
    idx = close.loc[start:end].index
    full = close.index
    pos0 = full.get_loc(idx[0])

    reb = cfg['rebal']
    reb_dates, prev = set(), None
    for d in idx:
        key = (d.year, {'Q': (d.month - 1) // 3,
                        'M': d.month,
                        '2M': (d.month - 1) // 2}[reb])
        if prev is None or key != prev:
            reb_dates.add(d)
        prev = key

    w = pd.Series(dtype=float)
    nav = capital
    navs, dates = [], []
    review_px = {}
    for i in range(pos0, pos0 + len(idx)):
        t = full[i]
        if len(w):
            r = rets.iloc[i].reindex(w.index).fillna(0.0)
            gross = float((w * r).sum())
            nav *= (1 + gross)
            w = (w * (1 + r)) / (1 + gross)

        sl = cfg.get('stop_loss')
        if sl is not None and len(w):
            tdy = close.iloc[i]
            for s_ in list(w.index):
                rp = review_px.get(s_)
                if rp is not None and not np.isnan(tdy[s_]) and tdy[s_] / rp - 1 <= sl:
                    w = w.drop(labels=s_)

        if t in reb_dates:
            tw = pd.Series(sel(close, t))
            old = w.reindex(tw.index.union(w.index)).fillna(0.0)
            new = tw.reindex(old.index).fillna(0.0)
            turn = float((new - old).abs().sum())
            nav *= (1 - cost * turn)
            w = new
            review_px = {s_: close.iloc[i][s_] for s_ in w.index}

        navs.append(nav)
        dates.append(t)

    nav_s = pd.Series(navs, index=dates)
    r_p = nav_s.pct_change().dropna()
    yrs = max((dates[-1] - dates[0]).days / 365.25, 0.01)
    tot = nav_s.iloc[-1] / capital - 1
    cagr = (1 + tot) ** (1 / yrs) - 1
    dd = (nav_s / nav_s.cummax() - 1).min()
    sd = r_p.std() * np.sqrt(252)
    return {'nav': nav_s, 'final': nav_s.iloc[-1], 'pnl': nav_s.iloc[-1] - capital,
            'total_%': tot * 100, 'cagr_%': cagr * 100, 'mdd_%': dd * 100,
            'sharpe': (cagr / sd) if sd > 0 else 0.0}


# =====================================================================
#  PERFORMANCE ENGINE (contest metrics)
# =====================================================================
def safe_float(v):
    try:
        if isinstance(v, pd.Series):
            v = v.iloc[0]
        return float(v)
    except Exception:
        return 0.0


class PerformanceEngine:
    def __init__(self, nav_series, trades, capital0, start, end):
        self.nav = pd.Series(nav_series).sort_index()
        self.trades = pd.DataFrame(trades) if trades else pd.DataFrame()
        self.cap0 = float(capital0)
        self.start, self.end = start, end

    def metrics(self):
        nav = self.nav
        rets = nav.pct_change().dropna()
        yrs = max((nav.index[-1] - nav.index[0]).days / 365.25, 0.01)
        tot = nav.iloc[-1] / self.cap0 - 1
        cagr = (1 + tot) ** (1 / yrs) - 1
        cum = (1 + rets).cumprod()
        dd = (cum / cum.expanding().max() - 1)
        mdd = dd.min()
        sd = rets.std() * np.sqrt(252)
        sharpe = cagr / sd if sd > 0 else 0.0

        # trade-level stats from episodes (realized P&L per symbol round-trip)
        acc = glr = np.nan
        n_trades = int(len(self.trades))
        wins = losses = []
        realized = {}
        if not self.trades.empty:
            # rebuild episode P&L per symbol (flat->flat round trips)
            by_sym = {}
            for _, r in self.trades.sort_values('date').iterrows():
                b = by_sym.setdefault(r['symbol'], {'sh': 0, 'cost': 0.0, 'pnl': []})
                if r['action'] == 'BUY':
                    b['sh'] += r['shares']
                    b['cost'] += r['gross'] + r['cost']
                else:
                    if b['sh'] > 0:
                        avg = b['cost'] / b['sh']
                        realized_amt = (r['gross'] - r['cost']) - avg * r['shares']
                        b['pnl'].append(realized_amt)
                        b['cost'] -= avg * r['shares']
                        b['sh'] -= r['shares']
                        if b['sh'] == 0:
                            realized[r['symbol']] = sum(b['pnl'])
                            b['pnl'] = []
            if realized:
                v = np.array(list(realized.values()))
                wins = v[v > 0]
                losses = v[v < 0]
                acc = len(wins) / len(v) * 100
                glr = (wins.mean() / abs(losses.mean())) if len(losses) and losses.mean() != 0 else np.inf

        turnover_val = safe_float(self.trades['gross'].sum()) if n_trades else 0.0
        costs = safe_float(self.trades['cost'].sum()) if n_trades else 0.0
        return {
            'Final Portfolio Value (Rs)': safe_float(nav.iloc[-1]),
            'Total Net PNL (Rs)': safe_float(nav.iloc[-1]) - self.cap0,
            'Total Return (%)': tot * 100,
            'Annualised Return (%)': cagr * 100,
            'Maximum Drawdown (%)': mdd * 100,
            'Sharpe Ratio': sharpe,
            'Gain/Loss Ratio': glr,
            'Accuracy (%)': acc,
            'Total Trades': n_trades,
            'Turnover Value (Rs)': turnover_val,
            'Total Transaction Costs (Rs)': costs,
            '_nav': nav, '_rets': rets, '_dd': dd,
        }

    def print_report(self, bench_close=None):
        m = self.metrics()
        sep = '=' * 66
        print(f"\n{sep}\n  PORTFOLIO PERFORMANCE REPORT  "
              f"({self.start} -> {self.end})\n{sep}")
        for k, v in m.items():
            if k.startswith('_'):
                continue
            if 'Rs' in k:
                print(f"  {k:<34} Rs {v:>16,.0f}")
            elif '%' in k or 'Return' in k or 'Drawdown' in k or 'Accuracy' in k:
                print(f"  {k:<34} {v:>17.2f} %")
            else:
                print(f"  {k:<34} {v:>17.2f}")
        if bench_close is not None:
            b = bench_close.reindex(self.nav.index, method='ffill').dropna()
            btot = b.iloc[-1] / b.iloc[0] - 1
            yrs = (b.index[-1] - b.index[0]).days / 365.25
            bcagr = (1 + btot) ** (1 / yrs) - 1
            br = b.pct_change().dropna()
            bcum = (1 + br).cumprod()
            bmdd = (bcum / bcum.expanding().max() - 1).min()
            bsh = bcagr / (br.std() * np.sqrt(252)) if br.std() > 0 else 0
            print(f"\n{sep}\n  BENCHMARK\n{sep}")
            print(f"  {'Total Return (%)':<34} {btot*100:>17.2f} %")
            print(f"  {'Annualised Return (%)':<34} {bcagr*100:>17.2f} %")
            print(f"  {'Maximum Drawdown (%)':<34} {bmdd*100:>17.2f} %")
            print(f"  {'Sharpe Ratio':<34} {bsh:>17.2f}")
            print(f"\n  Alpha (annualised) : {m['Annualised Return (%)'] - bcagr*100:>8.2f} %")
        return m


# =====================================================================
#  PLOTS
# =====================================================================
def plot_results(nav_s, rets, dd, bench, metrics, fname='portfolio_performance.png'):
    nav = pd.Series(nav_s).sort_index()
    fig, axes = plt.subplots(3, 2, figsize=(17, 13))
    fig.suptitle('BMOM 12-1 - Buffered Momentum | Performance Analysis',
                 fontsize=15, fontweight='bold')

    ax = axes[0, 0]
    ax.plot(nav.index, nav / 1e7, lw=2, color='steelblue', label='Portfolio')
    for nm, s in bench.items():
        s = s.reindex(nav.index, method='ffill')
        ax.plot(s.index, s / s.iloc[0], lw=1.4, alpha=0.85, label=nm)
    ax.axhline(1, color='grey', ls='--', alpha=0.6)
    ax.set_title('NAV (₹ Cr) vs Benchmarks (rebased)')
    ax.legend(fontsize=9); ax.grid(True, alpha=0.3)

    ax = axes[0, 1]
    cp = (1 + rets).cumprod() - 1
    ax.plot(cp.index, cp * 100, lw=2, color='green', label='Portfolio')
    for nm, s in bench.items():
        b = s.reindex(cp.index, method='ffill').pct_change().fillna(0)
        cb = (1 + b).cumprod() - 1
        ax.plot(cb.index, cb * 100, lw=1.4, alpha=0.8, label=nm)
    ax.set_title('Cumulative Return (%)'); ax.legend(fontsize=9)
    ax.grid(True, alpha=0.3)

    ax = axes[1, 0]
    ax.fill_between(dd.index, dd * 100, 0, alpha=0.35, color='red')
    ax.plot(dd.index, dd * 100, lw=0.8, color='darkred')
    ax.set_title(f"Drawdown (Max {metrics['Maximum Drawdown (%)']:.1f}%)")
    ax.grid(True, alpha=0.3)

    ax = axes[1, 1]
    try:
        monthly = rets.resample('ME').apply(lambda x: (1 + x).prod() - 1) * 100
        piv = monthly.groupby([monthly.index.year, monthly.index.month]).first().unstack()
        piv.columns = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul',
                       'Aug', 'Sep', 'Oct', 'Nov', 'Dec'][:len(piv.columns)]
        sns.heatmap(piv, annot=True, fmt='.0f', cmap='RdYlGn', center=0,
                    ax=ax, linewidths=0.4, cbar_kws={'label': '%'})
        ax.set_title('Monthly Returns (%)')
    except Exception as e:
        ax.text(0.5, 0.5, f'heatmap err: {e}', ha='center', va='center')

    ax = axes[2, 0]
    r = rets.dropna() * 100
    ax.hist(r, bins=60, color='steelblue', edgecolor='white')
    ax.axvline(r.mean(), color='red', ls='--', lw=2,
               label=f'Mean {r.mean():.2f}%')
    ax.set_title('Daily Return Distribution'); ax.legend()

    ax = axes[2, 1]; ax.axis('off')
    txt = "\n".join([
        "  KEY METRICS", "  " + "-" * 40,
        f"  Total PNL          : ₹ {metrics['Total Net PNL (Rs)']/1e7:>7.2f} Cr",
        f"  Final Value        : ₹ {metrics['Final Portfolio Value (Rs)']/1e7:>7.2f} Cr",
        f"  Annualised Return  : {metrics['Annualised Return (%)']:>10.2f} %",
        f"  Max Drawdown       : {metrics['Maximum Drawdown (%)']:>10.2f} %",
        f"  Sharpe Ratio       : {metrics['Sharpe Ratio']:>10.2f}",
        "  " + "-" * 40,
        f"  Accuracy (round trips): {metrics['Accuracy (%)']:>7.1f} %",
        f"  Gain/Loss Ratio    : {metrics['Gain/Loss Ratio']:>10.2f}",
        f"  Total Trades       : {int(metrics['Total Trades']):>10d}",
        f"  Txn Costs Paid     : ₹ {metrics['Total Transaction Costs (Rs)']/1e5:>7.1f} L",
    ])
    ax.text(0.02, 0.98, txt, va='top', fontsize=11, family='monospace',
            transform=ax.transAxes,
            bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.6))

    plt.tight_layout()
    plt.savefig(fname, dpi=200, bbox_inches='tight')
    plt.close()
    print(f'  chart saved -> {fname}')


# =====================================================================
#  STRESS MODULE 1 : FROZEN OUT-OF-SAMPLE FORWARD TEST (H1-2026)
# =====================================================================
def oos_forward_test(close, is_start=IS_START, is_end=IS_END,
                     oos_start=OOS_START, oos_end=OOS_END):
    print('\n' + '=' * 70)
    print('  STRESS 1 : FROZEN OUT-OF-SAMPLE TEST  (Jan-Jun 2026)')
    print('=' * 70)
    is_r = fast_backtest(close, is_start, is_end)
    oos_r = fast_backtest(close, oos_start, oos_end)
    tbl = pd.DataFrame({
        'In-Sample 21-25': [is_r['pnl'] / 1e7, is_r['cagr_%'], is_r['mdd_%'], is_r['sharpe']],
        'Out-of-Sample H1-26': [oos_r['pnl'] / 1e7, oos_r['cagr_%'],
                                oos_r['mdd_%'], oos_r['sharpe']],
    }, index=['Net PNL (Cr)', 'CAGR %', 'Max DD %', 'Sharpe'])
    print('\n', tbl.round(2).to_string())
    deg = oos_r['sharpe'] / is_r['sharpe'] if is_r['sharpe'] else np.nan
    print(f"\n  OOS return {oos_r['total_%']:+.1f}%  vs  Nifty 50 H1-26 ~ -8.7%")
    print(f"  Sharpe degradation ratio : {deg:.2f}  (>0.7 generalises well)")
    return {'is': is_r, 'oos': oos_r, 'degradation': deg}


# =====================================================================
#  STRESS MODULE 2 : CRISIS REPLAY (windows inside 2021-26)
# =====================================================================
CRISIS_WINDOWS = {
    '2022 rate-hike selloff': ('2022-01-17', '2022-06-17'),
    'Adani-Hindenburg':       ('2023-01-24', '2023-03-20'),
    'Election-day shock':     ('2024-06-03', '2024-06-05'),
    'FII exit / SMID crash':  ('2024-09-26', '2025-03-03'),
    'H1-2026 correction':     ('2026-01-01', '2026-06-30'),
}

def crisis_replay(nav, nifty):
    print('\n' + '=' * 70)
    print('  STRESS 2 : CRISIS REPLAY')
    print('=' * 70)
    rows = []
    for name, (s, e) in CRISIS_WINDOWS.items():
        s, e = pd.to_datetime(s), pd.to_datetime(e)
        seg = nav.loc[(nav.index >= s) & (nav.index <= e)]
        bseg = nifty.loc[(nifty.index >= s) & (nifty.index <= e)]
        if len(seg) < 2:
            continue
        p = safe_float(seg.iloc[-1] / seg.iloc[0] - 1) * 100
        b = safe_float(bseg.iloc[-1] / bseg.iloc[0] - 1) * 100 if len(bseg) > 1 else np.nan
        dd = safe_float(((seg / seg.expanding().max() - 1).min())) * 100
        rows.append([name, p, b, p - b, dd])
    df = pd.DataFrame(rows, columns=['Crisis', 'Portfolio %', 'Nifty50 %',
                                     'Excess %', 'Port MaxDD %'])
    print('\n', df.round(2).to_string(index=False))
    print(f"\n  Outperformed in {int((df['Excess %'] > 0).sum())}/{len(df)} crises")
    return df


# =====================================================================
#  STRESS MODULE 3 : PARAMETER SENSITIVITY (neighbourhood plateau)
# =====================================================================
def sensitivity_grid(close):
    print('\n' + '=' * 70)
    print('  STRESS 3 : PARAMETER SENSITIVITY')
    print('=' * 70)
    rows = []
    base = dict(CONFIG)
    variants = []
    for keep in [18, 20, 25, 30]:
        for n in [6, 7, 8, 10]:
            for stop in [None, -0.15, -0.25]:
                c = dict(base, keep_rank=keep, n_stocks=n, stop_loss=stop)
                variants.append(c)
    variants += [dict(base, rebal='M'), dict(base, rebal='2M'),
                 dict(base, mom_skip=10), dict(base, mom_skip=42),
                 dict(base, mom_lb=189), dict(base, ind_cap=2),
                 dict(base, ind_cap=None), dict(base, stop_loss=None),
                 dict(base, stop_loss=-0.15), dict(base, max_w=0.15),
                 dict(base, max_w=0.25), dict(base, vol_window=40),
                 dict(base, vol_window=90), dict(base, mom_lb=126)]
    print(f'  running {len(variants)} neighbouring configs …')
    for i, c in enumerate(variants, 1):
        try:
            r = fast_backtest(close, IS_START, IS_END, cfg=c)
            ro = fast_backtest(close, OOS_START, OOS_END, cfg=c)
            rows.append({'src': 'grid', 'n': c['n_stocks'], 'keep': c['keep_rank'],
                         'stop': c['stop_loss'], 'rebal': c['rebal'],
                         'skip': c['mom_skip'], 'lb': c['mom_lb'],
                         'IS_PNL_Cr': r['pnl'] / 1e7, 'IS_MDD': r['mdd_%'],
                         'IS_Sharpe': r['sharpe'], 'OOS_%': ro['total_%']})
        except Exception:
            pass
        for r_ in rows:
            r_.setdefault('src', 'variant')
        if i % 10 == 0:
            print(f'    {i}/{len(variants)}')
    df = pd.DataFrame(rows)
    print('\n  IS PNL (Cr) distribution across all neighbours:')
    print(df['IS_PNL_Cr'].describe().loc[['min', '25%', '50%', '75%', 'max']].round(2).to_string())
    print(f"\n  Configs ≥ ₹5 Cr PNL : {(df['IS_PNL_Cr'] >= 5).mean()*100:.0f}%")
    print(f"  Configs OOS-positive: {(df['OOS_%'] > 0).mean()*100:.0f}%")
    df.to_csv('stress_sensitivity_grid.csv', index=False)

    # plateau heatmap: keep x n at stop=-0.25
    try:
        sub = df[(df['src'] == 'grid') & (df['stop'] == -0.25)]
        piv = sub.pivot_table(index='keep', columns='n', values='IS_PNL_Cr')
        plt.figure(figsize=(7, 4.5))
        sns.heatmap(piv, annot=True, fmt='.1f', cmap='RdYlGn', center=8)
        plt.title('IS PNL (₹ Cr): buffer x portfolio size  (stop -25%)')
        plt.tight_layout(); plt.savefig('stress_sensitivity.png', dpi=200)
        plt.close()
        print('  saved -> stress_sensitivity.png')
    except Exception as e:
        print(f'  heatmap skipped: {e}')
    return df


# =====================================================================
#  STRESS MODULE 4 : MONTE CARLO BLOCK BOOTSTRAP
# =====================================================================
def monte_carlo(rets, capital=CAPITAL, n_sims=3000, block=21):
    print('\n' + '=' * 70)
    print('  STRESS 4 : MONTE CARLO BLOCK BOOTSTRAP')
    print('=' * 70)
    r = rets.dropna().values
    n = len(r)
    nb = int(np.ceil(n / block))
    starts = np.random.randint(0, max(n - block, 1), (n_sims, nb))
    segs = [r[s:s + block] for row in starts for s in row]
    mat = np.concatenate(segs).reshape(n_sims, nb * block)[:, :n]
    eq = capital * np.cumprod(1 + mat, axis=1)
    term = eq[:, -1]
    peak = np.maximum.accumulate(eq, axis=1)
    mdd = ((eq - peak) / peak).min(axis=1) * 100
    res = {'Median terminal (Cr)': np.median(term) / 1e7,
           'P5 terminal (Cr)': np.percentile(term, 5) / 1e7,
           'P95 terminal (Cr)': np.percentile(term, 95) / 1e7,
           'P(loss) %': float((term < capital).mean() * 100),
           'Median MDD %': float(np.median(mdd)),
           'P95 worst MDD %': float(np.percentile(mdd, 5))}
    for k, v in res.items():
        print(f'  {k:<24} {v:>8.2f}')
    plt.figure(figsize=(11, 5))
    plt.plot(eq[:300].T / 1e7, color='steelblue', alpha=0.03)
    plt.plot(np.median(eq, axis=0) / 1e7, color='darkred', lw=2.5, label='Median')
    plt.axhline(1, color='k', ls='--', alpha=.6, label='₹1 Cr')
    plt.title('Monte Carlo Fan - Block Bootstrap'); plt.ylabel('₹ Cr')
    plt.legend(); plt.tight_layout()
    plt.savefig('stress_montecarlo.png', dpi=200); plt.close()
    print('  saved -> stress_montecarlo.png')
    return res


# =====================================================================
#  STRESS MODULE 5 : TRANSACTION-COST SENSITIVITY
# =====================================================================
def cost_stress(close):
    print('\n' + '=' * 70)
    print('  STRESS 5 : TRANSACTION-COST SENSITIVITY')
    print('=' * 70)
    rows = []
    for c in [0.001, 0.0025, 0.005, 0.0075, 0.010, 0.015]:
        r = fast_backtest(close, IS_START, IS_END, cost=c)
        rows.append({'Cost/side %': c * 100, 'PNL (Cr)': r['pnl'] / 1e7,
                     'CAGR %': r['cagr_%'], 'Sharpe': r['sharpe']})
    df = pd.DataFrame(rows)
    print('\n', df.round(2).to_string(index=False))
    prof = df[df['PNL (Cr)'] > 0]
    if len(prof):
        print(f"\n  Alpha survives up to ~{prof['Cost/side %'].max():.2f}% per side")
    return df


# =====================================================================
#  STRESS MODULE 6 : REGIME & CAPTURE
# =====================================================================
def regime_analysis(rets, nifty_close):
    print('\n' + '=' * 70)
    print('  STRESS 6 : REGIME & CAPTURE ANALYSIS')
    print('=' * 70)
    nifty = nifty_close.reindex(rets.index).ffill()
    dma = nifty.rolling(200).mean()
    bull = (nifty > dma).reindex(rets.index).ffill()
    nret = nifty.pct_change().reindex(rets.index)
    rows = []
    for name, mask in [('Bull (Nifty>200DMA)', bull == True),
                       ('Bear (Nifty<200DMA)', bull == False)]:
        pr = rets[mask].dropna()
        br = nret[mask].dropna()
        if len(pr) < 5:
            continue
        rows.append({'Regime': name, 'Days': len(pr),
                     'Port Ann %': ((1 + pr).prod() ** (252 / len(pr)) - 1) * 100,
                     'Nifty Ann %': ((1 + br).prod() ** (252 / len(br)) - 1) * 100 if len(br) else np.nan,
                     'Hit Rate %': (pr > 0).mean() * 100})
    df = pd.DataFrame(rows)
    print('\n', df.round(2).to_string(index=False))
    al = pd.concat([rets.rename('p'), nret.rename('b')], axis=1).dropna()
    dn = al.nsmallest(20, 'b'); up = al.nlargest(20, 'b')
    d_cap = dn['p'].mean() / dn['b'].mean() if dn['b'].mean() else np.nan
    u_cap = up['p'].mean() / up['b'].mean() if up['b'].mean() else np.nan
    print(f"\n  Downside capture : {d_cap:.2f}   Upside capture : {u_cap:.2f}")
    print(f"  (low downside + high upside = convex profile)")
    return df


# =====================================================================
#  STRESS MODULE 7 : TAIL RISK + DSR + PERMUTATION
# =====================================================================
def tail_risk(rets, nav):
    print('\n' + '=' * 70)
    print('  STRESS 7A : TAIL-RISK METRICS')
    print('=' * 70)
    r = rets.dropna()
    var95 = np.percentile(r, 5); cvar95 = r[r <= var95].mean()
    var99 = np.percentile(r, 1); cvar99 = r[r <= var99].mean()
    dd = nav / nav.cummax() - 1
    ulcer = np.sqrt((dd ** 2).mean()) * 100
    yrs = (nav.index[-1] - nav.index[0]).days / 365.25
    cagr = (nav.iloc[-1] / nav.iloc[0]) ** (1 / yrs) - 1
    ds_std = r[r < 0].std() * np.sqrt(252)
    out = {'Daily VaR95 %': var95 * 100, 'Daily CVaR95 %': cvar95 * 100,
           'Daily VaR99 %': var99 * 100, 'Daily CVaR99 %': cvar99 * 100,
           'Skewness': stats.skew(r), 'Excess Kurtosis': stats.kurtosis(r),
           'Ulcer Index %': ulcer,
           'Calmar': cagr / abs(dd.min()) if dd.min() else np.nan,
           'Sortino': cagr / ds_std if ds_std else np.nan}
    for k, v in out.items():
        print(f'  {k:<20} {v:>8.2f}')
    return out


def deflated_sharpe(rets, n_trials):
    print('\n' + '=' * 70)
    print('  STRESS 7B : DEFLATED / PROBABILISTIC SHARPE')
    print('=' * 70)
    r = rets.dropna().values
    n = len(r)
    sr = r.mean() / r.std()
    g3 = float(stats.skew(r)); g4 = float(stats.kurtosis(r, fisher=False))
    var_sr = (1 - g3 * sr + ((g4 - 1) / 4) * sr ** 2) / (n - 1)
    gamma = 0.5772156649
    e1 = stats.norm.ppf(1 - 1 / max(n_trials, 2))
    e2 = stats.norm.ppf(1 - 1 / (max(n_trials, 2) * np.e))
    sr_star = np.sqrt(max(var_sr, 1e-12)) * ((1 - gamma) * e1 + gamma * e2)
    num = (sr - sr_star) * np.sqrt(n - 1)
    den = np.sqrt(max(1 - g3 * sr + ((g4 - 1) / 4) * sr ** 2, 1e-12))
    dsr = float(stats.norm.cdf(num / den))
    psr = float(stats.norm.cdf(sr * np.sqrt(n - 1) / den))
    print(f'  Sharpe (ann.)      : {sr*np.sqrt(252):.2f}')
    print(f'  Trials (N)         : {n_trials}')
    print(f'  PSR                : {psr:.2%}')
    print(f'  DSR                : {dsr:.2%}   (>95% survives selection bias)')
    return {'PSR': psr, 'DSR': dsr}


def permutation_test(close, actual_pnl, n_perm=40):
    print('\n' + '=' * 70)
    print('  STRESS 7C : PERMUTATION TEST (random-selection null)')
    print('=' * 70)
    cols = list(close.columns)
    n = CONFIG['n_stocks']
    pnls = []
    for i in range(n_perm):
        picks = list(np.random.choice(cols, n, replace=False))
        # random equal-weight buy-and-hold basket (same window, no costs spared)
        rets = close.pct_change(fill_method=None)
        idx = close.loc[IS_START:IS_END].index
        full = close.index; pos0 = full.get_loc(idx[0])
        reb_dates, prev = set(), None
        for d in idx:
            key = (d.year, (d.month - 1) // 3)
            if prev is not None and key != prev:
                reb_dates.add(d)
            prev = key
        w0 = pd.Series({p: 1 / n for p in picks})
        nav = CAPITAL
        navs = []
        wi = w0.copy()
        started = False
        for j in range(pos0, pos0 + len(idx)):
            t = full[j]
            if len(wi):
                rr = rets.iloc[j].reindex(wi.index).fillna(0.0)
                g = float((wi * rr).sum())
                nav *= (1 + g)
                wi = (wi * (1 + rr)) / (1 + g)
            navs.append(nav)
        pnls.append(nav - CAPITAL)
        if (i + 1) % 10 == 0:
            print(f'    {i+1}/{n_perm}')
    pnls = np.array(pnls)
    pval = float((pnls >= actual_pnl).mean())
    print(f"\n  Strategy PNL      : ₹{actual_pnl/1e7:.2f} Cr")
    print(f"  Random median PNL : ₹{np.median(pnls)/1e7:.2f} Cr")
    print(f"  Empirical p-value : {pval:.4f}")
    plt.figure(figsize=(9, 4.5))
    plt.hist(pnls / 1e7, bins=20, color='lightgrey', edgecolor='k')
    plt.axvline(actual_pnl / 1e7, color='red', lw=2.5, label=f'p={pval:.3f}')
    plt.title('Permutation Test vs Random 7-Stock Baskets')
    plt.xlabel('PNL (₹ Cr)'); plt.legend(); plt.tight_layout()
    plt.savefig('stress_permutation.png', dpi=200); plt.close()
    return {'p_value': pval}


# =====================================================================
#  STRESS MODULE 8 : WALK-FORWARD (no re-fit - strategy has no fit params)
# =====================================================================
def walk_forward(close, n_folds=5):
    print('\n' + '=' * 70)
    print('  STRESS 8 : WALK-FORWARD (rolling 12-month windows)')
    print('=' * 70)
    folds = []
    starts = pd.date_range('2021-01-01', periods=n_folds + 1, freq='12MS')
    for i in range(len(starts) - 1):
        s = starts[i] + pd.DateOffset(years=1)
        e = s + pd.DateOffset(years=1) - pd.Timedelta(days=1)
        if e > pd.Timestamp(OOS_END):
            break
        r = fast_backtest(close, s.strftime('%Y-%m-%d'), e.strftime('%Y-%m-%d'))
        folds.append({'Window': f'{s.date()} -> {e.date()}',
                      'Return %': r['total_%'], 'Sharpe': r['sharpe'],
                      'MDD %': r['mdd_%']})
        print(f"  {folds[-1]['Window']}: {r['total_%']:+.1f}%  (MDD {r['mdd_%']:.0f}%)")
    df = pd.DataFrame(folds)
    print(f"\n  Positive windows : {(df['Return %'] > 0).sum()}/{len(df)}")
    return df


# =====================================================================
#  EXCEL EXPORT
# =====================================================================
def export_excel(metrics, nav, trades, bench, holdings_log, fname='submission_summary.xlsx'):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    hd = Font(bold=True, color='FFFFFF')
    fill = PatternFill('solid', fgColor='1F4E78')
    title_font = Font(bold=True, size=13)

    def prettify(ws, widths):
        for i, wd in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = wd
        ws.freeze_panes = 'A2'

    ws = wb.active; ws.title = 'Summary'
    ws['A1'] = 'Finesse x Citadel Portfolio Challenge - Round 2 | BMOM 12-1 | Team: <TEAM NAME>'
    ws['A1'].font = title_font
    ws.append([])
    ws.append(['Metric', 'Portfolio', 'Nifty 50'])
    for c in 'AB': ws[f'{c}3'].font = hd; ws[f'{c}3'].fill = fill
    b = bench.get('Nifty 50')
    bseg = b.reindex(nav.index, method='ffill').dropna() if b is not None else None
    b_tot = (bseg.iloc[-1] / bseg.iloc[0] - 1) * 100 if bseg is not None else np.nan
    yrs = (nav.index[-1] - nav.index[0]).days / 365.25
    b_ann = ((1 + b_tot / 100) ** (1 / yrs) - 1) * 100 if bseg is not None else np.nan
    pairs = [
        ('Starting Capital (Rs)', CAPITAL, ''),
        ('Final Portfolio Value (Rs)', metrics['Final Portfolio Value (Rs)'], ''),
        ('Total Net PNL (Rs)', metrics['Total Net PNL (Rs)'], ''),
        ('Total Return (%)', round(metrics['Total Return (%)'], 2), round(b_tot, 2)),
        ('Annualised Return (%)', round(metrics['Annualised Return (%)'], 2), round(b_ann, 2)),
        ('Maximum Drawdown (%)', round(metrics['Maximum Drawdown (%)'], 2), ''),
        ('Sharpe Ratio', round(metrics['Sharpe Ratio'], 2), ''),
        ('Gain/Loss Ratio', round(metrics['Gain/Loss Ratio'], 2)
            if np.isfinite(metrics['Gain/Loss Ratio']) else 'n/a'),
        ('Accuracy (%)', round(metrics['Accuracy (%)'], 2) if metrics['Accuracy (%)'] == metrics['Accuracy (%)'] else ''),
        ('Total Trades', metrics['Total Trades'], ''),
        ('Turnover Value (Rs)', round(metrics['Turnover Value (Rs)']), ''),
        ('Transaction Costs (Rs)', round(metrics['Total Transaction Costs (Rs)']), ''),
    ]
    for p in pairs:
        ws.append(list(p))
    prettify(ws, [34, 22, 14])
    ws.freeze_panes = 'A4'

    ws2 = wb.create_sheet('Quarterly_Holdings')
    ws2.append(['Review Date', 'Symbol', 'Weight %'])
    for c in 'ABC': ws2[f'{c}1'].font = hd; ws2[f'{c}1'].fill = fill
    for d, dct in holdings_log:
        for s, w in dct.items():
            ws2.append([str(d.date()), s.replace('.NS', ''), round(w * 100, 2)])
    prettify(ws2, [14, 14, 10])
    ws3 = wb.create_sheet('Trades')
    if len(trades):
        ws3.append(list(trades.columns))
        for _, r in trades.iterrows():
            ws3.append([str(r['date'].date()) if hasattr(r['date'], 'date') else str(r['date']),
                        r['symbol'].replace('.NS', ''), r['action'], r['reason'],
                        r['shares'], round(r['price'], 2), round(r['gross']), round(r['cost'])])
    prettify(ws3, [12, 12, 8, 12, 10, 10, 12, 10])
    ws4 = wb.create_sheet('Daily_NAV')
    ws4.append(['Date', 'Portfolio Value (Rs)'])
    for t, v in nav.items():
        ws4.append([str(t.date()), round(float(v))])
    prettify(ws4, [12, 20])
    wb.save(fname)
    print(f'  saved -> {fname}')


# =====================================================================
#  MAIN
# =====================================================================
def main():
    print('\n' + '#' * 70)
    print('#   BMOM 12-1  -  Buffered Momentum | full contest run')
    print('#' * 70)

    close = download_data()
    bench = download_benchmarks()
    n50 = bench.get('Nifty 50', pd.Series(dtype=float))

    # -- headline backtest (event-driven, contest accounting) --
    print('\n' + '=' * 70)
    print(f'  BASE BACKTEST  {IS_START} -> {IS_END}   (₹1 Cr, 0.1% txn cost)')
    print('=' * 70)
    bt = Backtester(close, IS_START, IS_END).run(verbose=True)

    pe = PerformanceEngine(bt.nav_series, bt.trades, CAPITAL, IS_START, IS_END)
    metrics = pe.print_report(bench_close=n50)
    rets, dd, nav = metrics['_rets'], metrics['_dd'], metrics['_nav']

    plot_results(bt.nav_series, rets, dd, bench, metrics)
    nav.to_csv('daily_nav.csv', header=['portfolio_value'])
    pd.DataFrame(bt.trades).to_csv('trade_history.csv', index=False)

    with open('backtest_results.json', 'w') as f:
        json.dump({k: (float(v) if isinstance(v, (int, float, np.floating)) else str(v))
                   for k, v in metrics.items() if not k.startswith('_')},
                  f, indent=4, default=str)

    # -- stress suite --
    oos = oos_forward_test(close)
    stitched_nav = fast_backtest(close, IS_START, OOS_END)['nav']
    crisis_df = crisis_replay(pd.concat([nav, stitched_nav.loc[nav.index[-1]:]]).sort_index(), n50)
    crisis_df.to_csv('stress_crisis_replay.csv', index=False)
    sens = sensitivity_grid(close)
    mc = monte_carlo(rets)
    costs = cost_stress(close)
    reg = regime_analysis(rets, n50)
    tail = tail_risk(rets, nav)
    dsr = deflated_sharpe(rets, n_trials=max(len(sens), 20))
    perm = permutation_test(close, metrics['Total Net PNL (Rs)'])
    wf = walk_forward(close)

    with open('stress_summary.json', 'w') as f:
        json.dump({'oos_degradation': oos['degradation'],
                   'permutation_p': perm['p_value'],
                   'PSR': dsr['PSR'], 'DSR': dsr['DSR'],
                   'monte_carlo': {k: float(v) for k, v in mc.items()},
                   'tail_risk': {k: float(v) for k, v in tail.items()}},
                  f, indent=4, default=str)

    # -- excel --
    export_excel(metrics, nav, pd.DataFrame(bt.trades), bench, bt.selector_log)

    print('\n' + '#' * 70)
    print('#   DONE - outputs:')
    print('#' * 70)
    for f_ in ['portfolio_performance.png', 'daily_nav.csv', 'trade_history.csv',
               'backtest_results.json', 'stress_sensitivity_grid.csv',
               'stress_sensitivity.png', 'stress_montecarlo.png',
               'stress_permutation.png', 'stress_crisis_replay.csv',
               'stress_summary.json', 'submission_summary.xlsx']:
        print(f'   {f_}')


if __name__ == '__main__':
    main()
